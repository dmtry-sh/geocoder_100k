import geocoder
from openpyxl import *
from proxy import *

try:
    doc_name = input('Введите название документа (без .xlsx): ')
    wb = load_workbook('исходники/{}.xlsx'.format(doc_name))
    ws = wb.active
    last = len(ws['B'])
except FileNotFoundError as err:
    print("Такой файл не найден. Попробуйте снова (файл должен лежать в папке 'исходники')")
    sleep = input('Нажмите любую клавишу, чтобы завершить работу программы')
    raise SystemExit


print('Получаю список прокси серверов')
proxy_list = Proxy()
print('Подключаюсь к серверу (может занять несколько минут)')
proxy = proxy_list.get_proxy()
print(proxy)
print('Запускаю парсер...')
for i in range(2, last+1):

    if (i-1) % 10 == 0:
        print('Сохраняю промежуточные результаты')
        wb.save('обработанные/{}_geo.xlsx'.format(doc_name))

    if i % 10000 == 0:
        print('Меняю сервер (может занять несколько минут)')
        proxy = proxy_list.next_server()

    print('Обработка {} адреса из {}'.format(i-1, last-1))
    address = ws['B{}'.format(i)].value
    url = 'https://geocode-maps.yandex.ru/1.x/?geocode={}'.format(address)

    try:
        r = requests.get(url, proxies={'http': proxy})
        soup = BeautifulSoup(r.content.decode('utf-8'), 'xml')
        try:
            pos = soup.find('pos').text.split(' ')
            pos = str(pos[1]) + ', ' + str(pos[0])
            ws['C{}'.format(i)] = pos
        except:
            ws['C{}'.format(i)] = ' '

    except requests.exceptions.ConnectionError:
        print('Ошибка сервера. Подключаюсь к новому')
        proxy = proxy_list.next_server()
        if proxy is None:
            proxy_list.update_proxy()
            proxy = proxy_list.get_proxy()
        r = requests.get(url, proxies={'http': proxy})
        soup = BeautifulSoup(r.content.decode('utf-8'), 'xml')
        pos = soup.find('pos').text.split(' ')
        pos = str(pos[1]) + ', ' + str(pos[0])
        ws['C{}'.format(i)] = pos

print('Сохраняю результаты в файл обработанные/{}_geo.xlsx'.format(doc_name))
wb.save('обработанные/{}_geo.xlsx'.format(doc_name))
sleep = input('Обработка успешно завершена. Нажмите любую клавишу, чтобы завершить работу.')

